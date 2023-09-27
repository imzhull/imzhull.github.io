# -*- coding: utf-8 -*-
'''
@文件    :excleFunc.py
@说明    :用来操作excle文件的方法
@版本    :1.0
'''

from openpyxl import load_workbook
from config import excle_path


def readExcle(sheetName):
    """
    @功能    :读取excle数据的方法
    @版本    :1.0
    """
    excle = load_workbook(excle_path)
    sheet = [row for row in excle[sheetName]]         # 把sheet表单的数据对象转换成数组
    keys = [key.value for key in sheet.pop(0)]      # 把第一行取出来并且读取具体的值并存放到keys数组中
    datas = []
    for row in sheet:
        values = []
        for data in row:
            data = data.value
            values.append(data)
        datas.append(dict(zip(keys,values)))        # 转换成字段并存到数组中
    return datas

def writeExcle(sheetName,cellName,value):
    """
    Excle表中写入数据
    """
    excle = load_workbook(excle_path)
    sheet = excle[sheetName]
    sheet[cellName].value = value
    excle.save(excle_path)




if __name__ == "__main__":
    readExcle()

    